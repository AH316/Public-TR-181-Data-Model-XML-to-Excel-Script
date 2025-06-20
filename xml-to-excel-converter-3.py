import os
import re
import fnmatch
import xml.etree.ElementTree as ET
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# --- Helper function to resolve dataType references recursively ---
def resolve_datatype_reference(datatype_name, xml_root, visited=None):
    """
    Recursively resolve a dataType reference, extracting type and size/range information.
    Walks up the <dataType> base chain until a concrete type (string, int, unsignedInt) is found,
    even if inheritance is multi-level.
    """
    if visited is None:
        visited = set()
    if datatype_name in visited:
        return None
    visited.add(datatype_name)

    # Find the datatype element with the given name
    dt_elem = None
    for dt in xml_root.findall(".//dataType"):
        if dt.get("name") == datatype_name:
            dt_elem = dt
            break
    if dt_elem is None:
        return None

    # Helper for extracting type and size/range/enumeration from primitive type elements
    def extract_type_info(elem):
        tag_name = elem.tag.lower().split('}')[-1]
        tag_name_norm = tag_name.lower()

        # Collect all <size> tags (for types like hexBinary, string, etc.)
        size_ranges = []
        for size_elem in elem.findall("size"):
            min_len = size_elem.get("minLength")
            max_len = size_elem.get("maxLength")
            if min_len and max_len:
                size_ranges.append(f"{min_len}:{max_len}")
            elif min_len:
                size_ranges.append(f"{min_len}:")
            elif max_len:
                size_ranges.append(f"{max_len}")

        # Handle <range> tag if present
        range_elem = elem.find("range")
        if range_elem is not None:
            min_val = range_elem.get("minInclusive")
            max_val = range_elem.get("maxInclusive")
            if min_val and max_val:
                size_ranges.append(f"{min_val}:{max_val}")
            elif min_val:
                size_ranges.append(f"{min_val}:")
            elif max_val:
                size_ranges.append(f"{max_val}")

        # Combine size/range into string with special handling for int, long, unsignedint
        if tag_name_norm in ["int", "long", "unsignedint"]:
            size_range_str = f"[{', '.join(size_ranges)}]" if size_ranges else ""
        else:
            size_range_str = f"({', '.join(size_ranges)})" if size_ranges else ""

        enums = elem.findall("enumeration")
        enum_values = [e.get("value") for e in enums if e.get("value") is not None]
        enum_str = ",".join(enum_values) if enum_values else None

        # Compose base type
        base = tag_name_norm
        # If enumeration present, treat as enum type only if string has other content
        if enum_values:
            # If the type is 'string' and all children are <enumeration>, show just 'string'
            if tag_name_norm == "string" and all(child.tag.lower().split('}')[-1] == "enumeration" for child in elem):
                base = "string"
                enum_values = []  # Suppress enum output for pure enum cases
            else:
                base = "enum"
        # Compose type string
        if size_ranges:
            type_str = f"{base}{size_range_str}"
        else:
            type_str = base
        if enum_values:
            type_str += f"[{enum_str}]"
        return type_str

    # Recursive base chain traversal to resolve ultimate type and size/range/enumeration
    def walk_base_chain(elem, xml_root, visited):
        # Look for any child that defines a type (primitive or otherwise)
        for child in elem:
            tag_name = child.tag.lower().split('}')[-1]
            # Ignore known non-type tags
            if tag_name in ['description']:
                continue
            # Apply extract_type_info to any child element
            type_info = extract_type_info(child)
            if type_info:
                return type_info
        # Fallback: look for primitive tags anywhere under elem
        for tag in ["string", "int", "unsignedInt", "unsignedLong", "hexBinary", "dateTime", "boolean", "list"]:
            syntax_elem = elem.find(".//" + tag)
            if syntax_elem is not None:
                type_info = extract_type_info(syntax_elem)
                if type_info:
                    return type_info
        # Also check for enumeration in this element
        enums = elem.findall("enumeration")
        if enums:
            enum_values = [e.get("value") for e in enums if e.get("value") is not None]
            enum_str = ",".join(enum_values)
            return f"enum[{enum_str}]"
        # Also check for <size> and <range> tags directly on this element, and combine
        size_ranges = []
        for size_elem in elem.findall("size"):
            min_len = size_elem.get("minLength")
            max_len = size_elem.get("maxLength")
            if min_len and max_len:
                size_ranges.append(f"{min_len}:{max_len}")
            elif min_len:
                size_ranges.append(f"{min_len}:")
            elif max_len:
                size_ranges.append(f"{max_len}")
        range_elem = elem.find("range")
        if range_elem is not None:
            min_val = range_elem.get("minInclusive")
            max_val = range_elem.get("maxInclusive")
            if min_val and max_val:
                size_ranges.append(f"{min_val}:{max_val}")
            elif min_val:
                size_ranges.append(f"{min_val}:")
            elif max_val:
                size_ranges.append(f"{max_val}")
        # Define tag_name_norm before using it below
        tag_name_norm = elem.tag.lower().split('}')[-1]
        if tag_name_norm in ["int", "long", "unsignedint"]:
            size_range_str = f"[{', '.join(size_ranges)}]" if size_ranges else ""
        else:
            size_range_str = f"({', '.join(size_ranges)})" if size_ranges else ""
        if size_range_str:
            return f"{tag_name_norm}{size_range_str}"
        # Walk up the base chain recursively
        base = elem.get("base")
        if base and base not in visited:
            visited.add(base)
            parent_elem = None
            for dt in xml_root.findall(".//dataType"):
                if dt.get("name") == base:
                    parent_elem = dt
                    break
            if parent_elem is not None:
                return walk_base_chain(parent_elem, xml_root, visited)
        return None

    # Try to resolve type recursively
    resolved_type = walk_base_chain(dt_elem, xml_root, visited)
    if resolved_type:
        return resolved_type

    # As fallback, try to infer type from any child that might define a type
    for child in dt_elem:
        tag_name = child.tag.lower().split('}')[-1]
        if tag_name not in ['description']:
            type_info = extract_type_info(child)
            if type_info:
                return type_info
    # As fallback, return None
    return None

def normalize_path(path):
    """Normalize a parameter path for comparison."""
    return path.lower().replace(" ", "").strip(".")

# Global macro substitution function
def substitute_macros(text, param_name=None, object_path=None):
    if not text:
        return ""

    def macro_replacer(match):
        macro = match.group(1).strip()

        if macro == "numentries":
            if param_name and param_name.endswith("NumberOfEntries"):
                table = param_name.replace("NumberOfEntries", "")
                full_table = f"{object_path}{table}" if object_path else table
                return f"The number of entries in the {full_table} table."
            return "The number of entries."

        if macro == "empty":
            return "an empty string"

        if macro == "pattern":
            return "a valid value matching the required pattern"

        if macro in ["reference", "referenceName", "noreference"]:
            return ""

        if macro.startswith("param|") or macro.startswith("object|") or macro.startswith("bibref|"):
            return macro.split("|", 1)[1]

        if macro.startswith("reference|"):
            content = macro.split("|", 1)[1]
            return content.replace("{{object}}", object_path.rstrip('.') if object_path else "this object")

        return ""

    try:
        result = re.sub(r"\{\{(.*?)\}\}", macro_replacer, text)
    except Exception as e:
        print(f"Error replacing macros in text: {text[:50]}... -> {e}")
        result = text
    return result.strip()

def clean_text(text):
    """Clean text by removing extra whitespace and newlines"""
    if text is None:
        return ""
    # Replace multiple whitespaces and newlines with a single space
    return re.sub(r'\s+', ' ', text).strip()

def extract_parameter_data(param_elem, parent_object_name, references_dict, templates_dict, html_descriptions, xml_root):
    """
    Extract data from a parameter element.
    
    Args:
        param_elem: The parameter XML element
        parent_object_name: Name of the parent object
        references_dict: Dictionary containing all references
        
    Returns:
        Dictionary containing parameter data
    """
    # Extract basic attributes
    name = param_elem.get('name', '')
    access = param_elem.get('access', '')
    version = param_elem.get('version', '')
    min_entries = param_elem.get('minEntries', '')
    max_entries = param_elem.get('maxEntries', '')
    ref = param_elem.get('ref', '')
    template_ref = param_elem.get('template', '')
    
    # Create sanitized full path (remove accidental double dots, trim leading/trailing dots)
    full_path = f"{parent_object_name.rstrip('.')}.{name}".replace(" ", "").strip(".").rstrip('.')
    # Initialize parameter data (remove specified keys)
    param_data = {
        'Object Name': parent_object_name,
        'Parameter Name': name,
        'Full Path': full_path,
        'Access': access,
        'Version': version,
        'Description': '',
        'Data Type': '',
        'Object Default': '',
        'Is Object': False
    }

    # Prefer HTML description lookup first (case-insensitive, normalized)
    full_path = param_data.get('Full Path', '')
    normalized = normalize_path(full_path)
    if normalized in html_descriptions:
        param_data['Description'] = html_descriptions[normalized]
    else:
        # Try relaxed matching by removing {i} placeholders
        relaxed = normalized.replace("{i}", "")
        found = False
        for key in html_descriptions.keys():
            if relaxed == key.replace("{i}", ""):
                param_data['Description'] = html_descriptions[key]
                found = True
                break
        if not found:
            # Fallback to XML embedded description
            found_xml_desc = False
            for desc_elem in param_elem:
                if 'description' in desc_elem.tag.lower():
                    desc_text = ''.join(desc_elem.itertext())
                    desc_text = substitute_macros(desc_text, param_name=name, object_path=parent_object_name)
                    param_data['Description'] = clean_text(desc_text)
                    found_xml_desc = True
            if not found_xml_desc:
                param_data['Description'] = "No HTML or XML description available"
    
    # Process syntax and data type, and extract Object Default if present
    for syntax_elem in param_elem:
        if 'syntax' in syntax_elem.tag.lower():
            # Always look for <dataType ref="..."> and resolve recursively
            is_list = False
            data_type = ''
            formatted_type = ''
            size_elem = None
            range_elem = None
            datatype_ref_found = None
            # Look for default value in syntax or its children
            default_found = False
            for type_elem in syntax_elem:
                tag_name = type_elem.tag.split('}')[-1].lower()
                if tag_name == 'default':
                    # Default value can be an attribute or text
                    value = type_elem.get('value')
                    if value is not None:
                        param_data['Object Default'] = value
                        default_found = True
                    elif type_elem.text:
                        param_data['Object Default'] = type_elem.text.strip()
                        default_found = True
                    continue
                if tag_name == 'datatype':
                    datatype_ref_found = type_elem.get('ref')
                    break
            if datatype_ref_found:
                resolved_type = resolve_datatype_reference(datatype_ref_found, xml_root)
                if resolved_type:
                    param_data['Data Type'] = resolved_type
                else:
                    param_data['Data Type'] = f"ref({datatype_ref_found})"
            else:
                # Old logic for list, string, int, unsignedInt, hexBinary, etc.
                for i, type_elem in enumerate(syntax_elem):
                    tag_name = type_elem.tag.split('}')[-1].lower()
                    if tag_name == 'default':
                        # Already handled above
                        continue
                    if tag_name == 'list':
                        is_list = True
                        size_elem = type_elem.find('.//size')
                        range_elem = type_elem.find('.//range')
                        # Also check if the list contains a hexbinary element
                        list_type_elem = None
                        for child in type_elem:
                            child_tag = child.tag.split('}')[-1].lower()
                            if child_tag == 'hexbinary':
                                list_type_elem = child
                                break
                        if list_type_elem is not None:
                            data_type = 'hexbinary'
                            # Aggregate all <size> tags for hexbinary list
                            if data_type == 'hexbinary':
                                sizes = list_type_elem.findall(".//size")
                                size_parts = []
                                for size in sizes:
                                    min_len = size.get('minLength')
                                    max_len = size.get('maxLength')
                                    if min_len and max_len:
                                        size_parts.append(f"{min_len}:{max_len}")
                                    elif min_len:
                                        size_parts.append(f"{min_len}:")
                                    elif max_len:
                                        size_parts.append(f"{max_len}")
                                formatted_type = f"hexbinary({', '.join(size_parts)})" if size_parts else "hexbinary"
                            else:
                                formatted_type = data_type
                            continue
                    elif tag_name in ['string', 'int', 'unsignedint', 'long', 'unsignedlong']:
                        data_type = tag_name
                        if not is_list:
                            size_elem = type_elem.find('.//size')
                            range_elem = type_elem.find('.//range')
                        # Try to find a <value> element for default
                        for value_elem in type_elem:
                            if 'value' in value_elem.tag.lower():
                                if value_elem.text:
                                    param_data['Object Default'] = value_elem.text.strip()
                                    default_found = True
                    elif tag_name == 'hexbinary':
                        data_type = tag_name
                        # Aggregate all <size> tags for hexbinary (non-list)
                        if data_type == 'hexbinary':
                            sizes = type_elem.findall(".//size")
                            size_parts = []
                            for size in sizes:
                                min_len = size.get('minLength')
                                max_len = size.get('maxLength')
                                if min_len and max_len:
                                    size_parts.append(f"{min_len}:{max_len}")
                                elif min_len:
                                    size_parts.append(f"{min_len}:")
                                elif max_len:
                                    size_parts.append(f"{max_len}")
                            formatted_type = f"hexbinary({', '.join(size_parts)})" if size_parts else "hexbinary"
                        else:
                            formatted_type = data_type
                        # Try to find a <value> element for default
                        for value_elem in type_elem:
                            if 'value' in value_elem.tag.lower():
                                if value_elem.text:
                                    param_data['Object Default'] = value_elem.text.strip()
                                    default_found = True
                    elif tag_name not in ['list', 'string', 'int', 'unsignedint', 'long', 'unsignedlong', 'datatype', 'hexbinary']:
                        data_type = tag_name
                        formatted_type = data_type
                        # Try to find a <value> element for default
                        for value_elem in type_elem:
                            if 'value' in value_elem.tag.lower():
                                if value_elem.text:
                                    param_data['Object Default'] = value_elem.text.strip()
                                    default_found = True
                if is_list:
                    if data_type == 'string':
                        if size_elem is not None:
                            min_len = size_elem.get('minLength')
                            max_len = size_elem.get('maxLength')
                            if min_len and max_len:
                                formatted_type = f"{data_type}({min_len}:{max_len})"
                            elif max_len:
                                formatted_type = f"{data_type}({max_len})"
                            else:
                                formatted_type = data_type
                        else:
                            formatted_type = data_type
                    elif data_type in ['int', 'unsignedint', 'long', 'unsignedlong']:
                        if range_elem is not None:
                            min_val = range_elem.get('minInclusive')
                            max_val = range_elem.get('maxInclusive')
                            if data_type == 'unsignedint':
                                # Use square brackets for unsignedint
                                if min_val and max_val:
                                    formatted_type = f"{data_type}[{min_val}:{max_val}]"
                                elif min_val:
                                    formatted_type = f"{data_type}[{min_val}:]"
                                elif max_val:
                                    formatted_type = f"{data_type}[:{max_val}]"
                                else:
                                    formatted_type = data_type
                            else:
                                if min_val and max_val:
                                    formatted_type = f"{data_type}({min_val}:{max_val})"
                                elif min_val:
                                    formatted_type = f"{data_type}({min_val}:)"
                                elif max_val:
                                    formatted_type = f"{data_type}(:{max_val})"
                                else:
                                    formatted_type = data_type
                        else:
                            formatted_type = data_type
                    elif data_type == 'hexbinary':
                        # formatted_type already set in loop above for hexbinary+list
                        if not formatted_type:
                            formatted_type = data_type
                    else:
                        formatted_type = data_type
                else:
                    if data_type == 'string':
                        if size_elem is not None:
                            min_len = size_elem.get('minLength')
                            max_len = size_elem.get('maxLength')
                            if min_len and max_len:
                                formatted_type = f"{data_type}({min_len}:{max_len})"
                            elif max_len:
                                formatted_type = f"{data_type}({max_len})"
                            else:
                                formatted_type = data_type
                        else:
                            formatted_type = data_type
                    elif data_type in ['int', 'unsignedint', 'long', 'unsignedlong']:
                        if range_elem is not None:
                            min_val = range_elem.get('minInclusive')
                            max_val = range_elem.get('maxInclusive')
                            if data_type in ['int', 'long', 'unsignedint']:
                                # Use square brackets for int, long, and unsignedint
                                if min_val and max_val:
                                    formatted_type = f"{data_type}[{min_val}:{max_val}]"
                                elif min_val:
                                    formatted_type = f"{data_type}[{min_val}:]"
                                elif max_val:
                                    formatted_type = f"{data_type}[:{max_val}]"
                                else:
                                    formatted_type = data_type
                            else:
                                formatted_type = data_type
                        else:
                            formatted_type = data_type
                    elif data_type == 'hexbinary':
                        # formatted_type already set in loop above for hexbinary
                        if not formatted_type:
                            formatted_type = data_type
                    else:
                        if not formatted_type:
                            formatted_type = data_type
                param_data['Data Type'] = formatted_type if formatted_type else data_type
    
    # Process template if present
    if template_ref:
        # Apply template data
        template_data = extract_template_data(template_ref, templates_dict)
        if template_data:
            # Only override if not already set
            if not param_data['Description'] and template_data.get('description'):
                param_data['Description'] = template_data['description']
            if not param_data['Data Type'] and template_data.get('data_type'):
                param_data['Data Type'] = template_data['data_type']
            if not param_data['Object Default'] and template_data.get('default_value'):
                param_data['Object Default'] = template_data['default_value']

    # Process reference if present
    if ref:
        # Resolve reference to actual value
        ref_data = resolve_reference(ref, references_dict)
        if ref_data:
            # Only override if not already set
            if not param_data['Description']:
                param_data['Description'] = ref_data.get('description', '')
            if not param_data['Data Type']:
                param_data['Data Type'] = ref_data.get('data_type', '')
            if not param_data['Object Default']:
                param_data['Object Default'] = ref_data.get('default_value', '')

    return param_data

def resolve_reference(ref_name, references_dict):
    """
    Resolve a reference to its actual data.
    
    Args:
        ref_name: Name of the reference to resolve
        references_dict: Dictionary containing all references
        
    Returns:
        Dictionary containing the resolved reference data
    """
    # Look up the reference in the references dictionary
    if ref_name in references_dict:
        ref_data = references_dict[ref_name]
        return {
            'description': f"Reference to: {ref_data}",
            'data_type': ''
        }
    return None

def extract_object_data(obj_elem, html_descriptions):
    """
    Extract and process data from an object XML element.
    
    Args:
        obj_elem: XML element representing an object
        html_descriptions: Dictionary mapping normalized full paths to HTML descriptions
        
    Returns:
        Dictionary containing all extracted object information
    """
    # Define namespace
    ns = {'dm': 'urn:broadband-forum-org:cwmp:datamodel-1-14'}
    
    # Extract basic object attributes
    name = obj_elem.get('name', '')
    access = obj_elem.get('access', '')
    min_entries = obj_elem.get('minEntries', '')
    max_entries = obj_elem.get('maxEntries', '')
    version = obj_elem.get('version', '')

    # Prefer HTML description lookup first (case-insensitive, normalized)
    full_path = name if name.endswith(".") else name + "."
    normalized = normalize_path(full_path)

    description = ""
    if normalized in html_descriptions:
        description = html_descriptions[normalized]
    else:
        desc_elem = obj_elem.find('dm:description', ns)
        description = clean_text(desc_elem.text) if desc_elem is not None and desc_elem.text else ""
    
    # Return structured object data
    return {
        'Object Name': name,
        'Access': access,
        'Min Entries': min_entries,
        'Max Entries': max_entries,
        'Version': version,
        'Description': description,
        'Is Object': True
    }

def extract_template_data(template_name, templates_dict):
    """
    Extract data from a template definition.
    
    Args:
        template_name: Name of the template to extract
        templates_dict: Dictionary containing all templates
        
    Returns:
        Dictionary containing template data
    """
    # Look up the template in the templates dictionary
    if template_name in templates_dict:
        template_data = templates_dict[template_name]
        # Only keep description and data_type (and default_value for param use)
        data = {
            'description': template_data.get('description', ''),
            'data_type': template_data.get('data_type', ''),
            'default_value': template_data.get('default_value', '')
        }
        # Check if this template inherits from another template
        parent_template = template_data.get('template', '')
        if parent_template:
            parent_data = extract_template_data(parent_template, templates_dict)
            if parent_data:
                if not data['description'] and parent_data.get('description'):
                    data['description'] = parent_data['description']
                if not data['data_type'] and parent_data.get('data_type'):
                    data['data_type'] = parent_data['data_type']
                if not data['default_value'] and parent_data.get('default_value'):
                    data['default_value'] = parent_data['default_value']
        return data
    return None

def extract_template_data_from_xml(template_elem):
    """
    Extract data from a template XML element.
    
    Args:
        template_elem: XML element representing a template
        
    Returns:
        Dictionary containing template data
    """
    # Extract basic template attributes
    name = template_elem.get('name', '')
    ref = template_elem.get('ref', '')
    template = template_elem.get('template', '')
    
    # Extract description
    description = ''
    for desc_elem in template_elem:
        if 'description' in desc_elem.tag.lower():
            description = clean_text(desc_elem.text) if desc_elem.text else ''
            break
    
    # Extract data type and default value
    data_type = ''
    default_value = ''
    for syntax_elem in template_elem:
        if 'syntax' in syntax_elem.tag.lower():
            # Get the first child element of syntax - this is the data type
            for type_elem in syntax_elem:
                data_type = type_elem.tag.split('}')[-1]  # Remove namespace
                
                # Check for value attribute
                value = type_elem.get('value', '')
                if value:
                    default_value = value
                
                # Check for default value in value element
                for value_elem in type_elem:
                    if 'value' in value_elem.tag.lower():
                        if value_elem.text:
                            default_value = value_elem.text
                break
    
    return {
        'name': name,
        'ref': ref,
        'template': template,
        'description': description,
        'data_type': data_type,
        'default_value': default_value
    }

def process_xml_file(xml_file_path, html_descriptions=None):
    """
    Process XML file to extract objects, parameters, templates, and references.
    
    Args:
        xml_file_path: Path to the XML file to process
        
    Returns:
        List of dictionaries containing all extracted data
    """
    all_data = []
    templates = {}
    references = {}
    if html_descriptions is None:
        html_descriptions = {}
    
    try:
        print(f"Processing XML file: {xml_file_path}")
        # Parse XML file
        tree = ET.parse(xml_file_path)
        root = tree.getroot()

        # Find the model element
        model = None
        for elem in root:
            if 'model' in elem.tag.lower():
                model = elem
                break

        if model is None:
            print("Error: Could not find model element in XML")
            return all_data

        # Extract templates and resolve inheritance
        # Extract all templates from XML (they are direct children of root)
        templates_found = []
        for elem in root:
            if 'template' in elem.tag.lower():
                templates_found.append(elem)

        # First pass: Extract all template data from XML
        for template in templates_found:
            template_data = extract_template_data_from_xml(template)
            if template_data['name']:
                templates[template_data['name']] = template_data

        # Second pass: Process template inheritance
        for template_name, template_data in templates.items():
            if template_data['template']:
                parent_template = template_data['template']
                if parent_template in templates:
                    parent_data = templates[parent_template]
                    # Only override if not already set
                    if not template_data['description'] and parent_data['description']:
                        template_data['description'] = parent_data['description']
                    if not template_data['data_type'] and parent_data['data_type']:
                        template_data['data_type'] = parent_data['data_type']
                    if not template_data['default_value'] and parent_data['default_value']:
                        template_data['default_value'] = parent_data['default_value']

        # references = { ref_name: targetParamRef }
        # Extract all references from XML (they are in the model element)
        references_found = []
        for elem in model.iter():
            if 'reference' in elem.tag.lower():
                references_found.append(elem)
        for ref in references_found:
            ref_name = ref.get('name', '')
            ref_target = ref.get('targetParamRef', '')
            if ref_name and ref_target:
                references[ref_name] = ref_target

        # Process all objects and their parameters (they are in the model element)
        objects_found = []
        for elem in model.iter():
            if 'object' in elem.tag.lower():
                objects_found.append(elem)

        print(f"Parsed {len(templates)} templates.")
        print(f"Parsed {len(references)} references.")
        print(f"Parsed {len(objects_found)} objects.")

        for obj_elem in objects_found:
            obj_data = extract_object_data(obj_elem, html_descriptions)
            # Clean up objects with empty or '.' names
            if not obj_data['Object Name'] or obj_data['Object Name'].strip() == ".":
                continue
            all_data.append(obj_data)

            # Process parameters within each object
            parameters_found = []
            for elem in obj_elem:
                if 'parameter' in elem.tag.lower():
                    parameters_found.append(elem)

            for param_elem in parameters_found:
                param_data = extract_parameter_data(param_elem, obj_data['Object Name'], references, templates, html_descriptions, root)
                # Clean up parameters with empty or '.' names or full paths
                if not param_data['Parameter Name'] or param_data['Parameter Name'].strip() == ".":
                    continue
                if not param_data['Full Path'] or param_data['Full Path'].strip() == ".":
                    continue
                all_data.append(param_data)

        print(f"Total entries processed: {len(all_data)}")

    except ET.ParseError as e:
        print(f"Error parsing XML: {e}")
    except Exception as e:
        print(f"Error processing XML: {e}")
        import traceback
        print(f"Full error traceback:\n{traceback.format_exc()}")

    return all_data

def create_excel(data, output_path):
    """
    Create Excel file from extracted data.
    
    Args:
        data: List of dictionaries containing the data to export
        output_path: Path where the Excel file should be saved
    """
    if not data:
        print("No data to export")
        return
    
    # Define column order for Excel output (remove specified columns)
    columns_order = [
        'Object Name', 'Parameter Name', 'Full Path',
        'Description', 'Data Type', 'Object Default', 'Is Object',
        'Access', 'Version'
    ]
    # Create DataFrame and save to Excel
    df = pd.DataFrame(data, columns=columns_order)
    
    try:
        df.to_excel(output_path, index=False, engine='openpyxl')
        print(f"Excel file created successfully: {output_path}")
        # Highlight object rows in Excel (light blue fill)
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill

        # Load workbook and access sheet
        wb = load_workbook(output_path)
        ws = wb.active

        # Define fill style for object rows
        object_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

        # Iterate over rows and apply fill to object rows
        is_object_col_idx = columns_order.index('Is Object') + 1
        for row_idx in range(2, ws.max_row + 1):  # Skip header row
            is_object = ws.cell(row=row_idx, column=is_object_col_idx).value
            if is_object:
                for col_idx in range(1, len(columns_order) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = object_fill

        # Save the updated workbook
        wb.save(output_path)
    except Exception as e:
        print(f"Error creating Excel file: {e}")

def main():
    # Get the directory where the script is located
    script_dir = os.path.dirname(os.path.abspath(__file__))

    from bs4 import NavigableString
    def extract_description(cells, desc_index):
        def convert_html_to_text(html_fragment):
            soup = BeautifulSoup(html_fragment, "html.parser")
            for br in soup.find_all(["br", "li", "p"]):
                br.insert_after("\n")
            text = soup.get_text()
            return re.sub(r'\n+', '\n', text).strip()

        if len(cells) > desc_index:
            raw_html = cells[desc_index].decode_contents()
        elif len(cells) >= 2:
            raw_html = cells[1].decode_contents()
        else:
            return ""

        return convert_html_to_text(raw_html)

    # Load HTML description lookup table from HTML file
    html_descriptions = {}
    html_path = None
    for root, _, files in os.walk(script_dir):
        for filename in fnmatch.filter(files, "tr-181*.html"):
            html_path = os.path.join(root, filename)
            break
        if html_path:
            break

    if html_path and os.path.exists(html_path):
        print(f"Found HTML file at: {html_path}")
        with open(html_path, 'r', encoding='utf-8') as f:
            soup = BeautifulSoup(f, 'html.parser')
        # New structure-aware HTML parsing logic
        table = soup.find("table", class_="data-model-table")
        if table:
            headers = [th.get_text(strip=True) for th in table.find_all("th")]
            desc_index = headers.index("Description") if "Description" in headers else 1
            rows = table.find_all("tr")
        else:
            rows = soup.find_all("tr")
            headers = []
            desc_index = 1

        last_object_path = ""

        for tr in rows:
            classes = tr.get("class", [])
            if not classes or ("object" not in classes and "parameter" not in classes):
                continue

            cells = tr.find_all("td")
            if len(cells) == 0:
                continue

            name = cells[0].get_text(strip=True)
            if not name:
                continue

            if name.lower().startswith("object definition") or name.lower().startswith("parameter definition"):
                continue

            description = extract_description(cells, desc_index)

            # Now process object or parameter assignment
            if "object" in classes:
                if not name.endswith("."):
                    name += "."
                last_object_path = name
                full_path = name
            else:
                if not last_object_path:
                    continue
                full_path = last_object_path + name

            if description:
                html_descriptions[normalize_path(full_path)] = description
    else:
        print("HTML file not found. Skipping HTML-based descriptions.")

    # File paths
    input_file = os.path.join(script_dir, "tr-181-2-19-0-cwmp-full.xml")
    output_file = os.path.join(script_dir, "TR-181-DM (Final).xlsx")

    # Process XML and create Excel
    print(f"Processing XML file: {input_file}")
    data = process_xml_file(input_file, html_descriptions=html_descriptions)
    print(f"Found {len(data)} entries")

    create_excel(data, output_file)

if __name__ == "__main__":
    main()